import numpy as np 
import pandas as pd

from sklearn.feature_extraction.text import CountVectorizer 
from sklearn.model_selection import train_test_split
from sklearn.naive_bayes import MultinomialNB

from sklearn.metrics import recall_score, precision_score, f1_score 
from sklearn.metrics import accuracy_score, confusion_matrix

# ===================================
import openpyxl 
from openpyxl import Workbook
# ===================================


file = "dataset.json"
data = pd.read_json(file)
# print(data.tail())
# print(data.shape)
data.sort_index(inplace = True)


labels = set()
for i in range(data.shape[0]):
	labels.add(data.iat[i,1])


vectorizer = CountVectorizer(stop_words='english', strip_accents='unicode')
all_features = vectorizer.fit_transform(data.tweets)
# print(all_features.shape)
# print(vectorizer.vocabulary_)
# print(all_features)
x_train, x_test, y_train, y_test = train_test_split(all_features, data.topic, test_size=0.3, random_state=88)

# print(x_train.shape)
# print(x_test.shape)

classifier = MultinomialNB()
classifier.fit(x_train,y_train)

# ==========
y_pred = classifier.predict(x_test)
acc_score = accuracy_score(y_test, y_pred)

# print(acc_score)
# print(classifier.labels())
conf_mat = confusion_matrix(y_test, y_pred, labels = list(labels))
print(conf_mat)
# ==========


num_correct = (y_test == classifier.predict(x_test)).sum()
num_incorrect = y_test.size -num_correct

# fraction = classifier.score(x_test, y_test) 
accuracy = num_incorrect/ (num_incorrect + num_correct)

recall = recall_score(y_test, classifier.predict(x_test),  average='macro')
precision = precision_score(y_test, classifier.predict(x_test),  average='macro')
f1_score = f1_score(y_test, classifier.predict(x_test),  average='macro')



print(f'{num_correct} classified correctly')
print(f'{num_incorrect} classified incorrectly')
print(f'The (testing) accuracy of the model is {1-accuracy:.2%}')
# print(f'The (testing) inaccuracy of the model is {fraction}')

print(f'Recall Score: {recall}')
print(f'Precision Score: {precision}')
print(f'f1 Score:  {f1_score}')

# =============================
#  CLASSIFICATION OF AN EXCEL FILE 
filename = "Dashboard.xlsx"

wb_obj = openpyxl.load_workbook(filename) 
sheet_obj = wb_obj.active 

row = sheet_obj.max_row
column = sheet_obj.max_column  

tweets = []
for i in range(3, row + 1): 
    cell_obj = sheet_obj.cell(row = i, column = 3) 
    tweets.append(cell_obj.value)

matrix = vectorizer.transform(tweets)
classification = classifier.predict(matrix)


# COMPUTING FOR THE PROBABILITIES OF THE LABELS
total = len(classification)
probabilities = {}
for i in labels:
	counter = 0
	for j in classification:
		if i == j:
			counter+=1
	probabilities[i] = str(round(counter/total,4) * 100) + '%'
print(probabilities)